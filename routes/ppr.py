import os
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import load_workbook
from .helpers import api_login_required
from db import get_ppr_list, get_ppr_departments, add_ppr_task, add_ppr_tasks_batch, close_ppr_task, update_ppr_close_date

ppr_bp = Blueprint('ppr', __name__, url_prefix='/api/ppr')

PPR_TEMPLATE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates_docs', 'ppr_template_3q.xlsx')


@ppr_bp.route('/list')
@api_login_required
def api_ppr_list():
    year = request.args.get('year', datetime.now().year, type=int)
    quarter = request.args.get('quarter', 1, type=int)
    department = request.args.get('department')
    tasks = get_ppr_list(year, quarter, department)
    return jsonify({"tasks": tasks})


@ppr_bp.route('/departments')
@api_login_required
def api_ppr_departments():
    year = request.args.get('year', datetime.now().year, type=int)
    quarter = request.args.get('quarter', 1, type=int)
    departments = get_ppr_departments(year, quarter)
    return jsonify({"departments": departments})


@ppr_bp.route('/close', methods=['POST'])
@api_login_required
def api_ppr_close():
    data = request.json or {}
    guid = data.get('guid')
    if not guid:
        return jsonify({'error': 'guid is required'}), 400
    comment = data.get('comment', '')
    latitude = data.get('latitude', 0.0)
    longitude = data.get('longitude', 0.0)
    close_date = data.get('close_date')
    success = close_ppr_task(guid, comment, latitude, longitude, close_date)
    if not success:
        return jsonify({'error': 'Task not found or already closed'}), 404
    return jsonify({'success': True})


@ppr_bp.route('/update-close-date', methods=['POST'])
@api_login_required
def api_ppr_update_close_date():
    data = request.json or {}
    guid = data.get('guid')
    close_date = data.get('close_date')
    if not guid or not close_date:
        return jsonify({'error': 'guid and close_date are required'}), 400
    success = update_ppr_close_date(guid, close_date)
    if not success:
        return jsonify({'error': 'Task not found or not closed'}), 404
    return jsonify({'success': True})


@ppr_bp.route('/add', methods=['POST'])
@api_login_required
def api_ppr_add():
    data = request.json or {}
    if "tasks" in data and isinstance(data["tasks"], list):
        if not data["tasks"]:
            return jsonify({'error': 'Empty tasks list'}), 400
        guids = add_ppr_tasks_batch(data["tasks"])
        return jsonify({"status": "ok", "count": len(guids), "guids": guids})
    if not data.get("number") or not data.get("name") or not data.get("name_department"):
        return jsonify({'error': 'number, name and name_department are required'}), 400
    guid = add_ppr_task(data)
    return jsonify({"status": "ok", "guid": guid})


@ppr_bp.route('/export-xlsx', methods=['POST'])
@api_login_required
def api_ppr_export_xlsx():
    data = request.json or {}
    numbers = data.get('numbers', [])
    date_str = data.get('date', '')

    if not numbers:
        return jsonify({'error': 'No PPR selected'}), 400
    if not date_str:
        return jsonify({'error': 'Date is required'}), 400

    if not os.path.exists(PPR_TEMPLATE):
        return jsonify({'error': 'Template not found'}), 500

    try:
        day, month, year = map(int, date_str.split('.'))
        date_obj = datetime(year, month, day)
    except (ValueError, IndexError):
        return jsonify({'error': 'Invalid date format. Expected DD.MM.YYYY'}), 400

    numbers_set = set(numbers)

    wb = load_workbook(PPR_TEMPLATE)
    ws = wb.active

    for row_idx in range(ws.max_row, 1, -1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            ws.delete_rows(row_idx)
            continue
        cell_str = str(cell_val).strip()
        if cell_str not in numbers_set:
            ws.delete_rows(row_idx)
        else:
            ws.cell(row=row_idx, column=3).value = date_obj
            ws.cell(row=row_idx, column=3).number_format = 'DD.MM.YYYY'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='ppr_3q.xlsx'
    )
